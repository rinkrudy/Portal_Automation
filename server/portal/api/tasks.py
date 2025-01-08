from celery import shared_task
import pandas as pd
import sys
import os
sys.path.append("C:/Users/User/Documents/Code/Projects/RPA/Portal/server/portal/api/module")
from test_workflow import Test_WorkFlow
from background_task import background
import pandas as pd
from .models import RequestDocument
from django.conf import settings
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def update_task_status(request_key, status):
    document = RequestDocument.objects.get(request_key = request_key)
    document.status = status
    document.save()

def update_task_result_path(request_key, result_path):
    document = RequestDocument.objects.get(request_key = request_key)
    document.path_result = result_path
    document.save()

@background(schedule=5)  # 5초 후 작업이 실행되도록 예약
def process_workflow(df_dict, request_key, background_mode):
    # DataFrame 복원
    
    df = pd.DataFrame.from_dict(df_dict)
    # Test_WorkFlow 작업 호출
    excel_buffer = io.BytesIO()
    df_result = Test_WorkFlow(df, request_key, background_mode)
    df_result.to_excel(excel_buffer, index=False)
    

    file_name = f"{request_key}.xlsx"

    file_path = os.path.join(settings.MEDIA_ROOT, file_name)

    # 엑셀 파일을 읽기
    workbook = load_workbook(excel_buffer)
    worksheet = workbook.active

    # 특정 컬럼에 조건을 적용해 스타일 설정
    fill_style_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 노란색 배경
    fill_style_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 노란색 배경
    fill_style_green = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")  # 노란색 배경

    # 예를 들어, 2번째 컬럼 (B 컬럼)에 값이 '특정값'이면 해당 셀의 배경색을 노란색으로 변경
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=4, max_col=5):
        for cell in row:
            if "불가" in str(cell.value):  # 셀 값에 '불가'가 포함되면
                cell.fill = fill_style_red
            elif "필요" in str(cell.value):
                cell.fill = fill_style_yellow
            elif "성공" in str(cell.value):
                cell.fill = fill_style_green

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)  # openpyxl로 저장
    excel_buffer.seek(0)  # 파일을 처음부터 읽을 수 있게 포인터 위치 설정

    file_content = ContentFile(excel_buffer.getvalue())
    saved_path = default_storage.save(file_name, file_content)
    download_url = os.path.join("http://localhost:8000/download/", file_name)

    update_task_status(request_key, 'COMPLETED')
    update_task_result_path(request_key, download_url)

